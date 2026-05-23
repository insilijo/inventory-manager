import csv
import io
import os
import json
import secrets
import sqlite3
import logging
from datetime import datetime
from contextlib import contextmanager
from typing import Optional

from dotenv import load_dotenv
load_dotenv()  # loads ./.env into the process env if present

import anthropic
from fastapi import FastAPI, Request, Form, HTTPException, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, PlainTextResponse, Response
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from twilio.request_validator import RequestValidator

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

DB_PATH = os.getenv("DB_PATH", "inventory.db")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
TWILIO_AUTH_TOKEN = os.getenv("TWILIO_AUTH_TOKEN", "")
VALIDATE_TWILIO = os.getenv("VALIDATE_TWILIO", "true").lower() == "true"
ADMIN_USER = os.getenv("ADMIN_USER", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")
SESSION_SECRET = os.getenv("SESSION_SECRET", "")


# ---------------------------------------------------------------------------
# Production env validation — fail fast at import time
# ---------------------------------------------------------------------------
# When VALIDATE_TWILIO=true we assume production: refuse to start with any of
# the credential env vars missing. Set VALIDATE_TWILIO=false for local dev.

if VALIDATE_TWILIO:
    # TWILIO_AUTH_TOKEN is optional — if absent, /sms returns 503 until configured.
    # ADMIN_PASSWORD is optional — if absent, admin endpoints are open (prototype mode).
    missing = [k for k, v in [
        ("ANTHROPIC_API_KEY", ANTHROPIC_API_KEY),
        ("SESSION_SECRET",    SESSION_SECRET),
    ] if not v]
    if missing:
        raise RuntimeError(
            "Production mode (VALIDATE_TWILIO=true) requires env vars: "
            + ", ".join(missing)
            + ". Set VALIDATE_TWILIO=false for local dev."
        )


app = FastAPI(title="FSFN Inventory")
app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET or secrets.token_urlsafe(32),
    session_cookie="fsfn_session",
    https_only=VALIDATE_TWILIO,  # only require HTTPS when in prod mode
    same_site="lax",
    max_age=60 * 60 * 12,  # 12 hours
)
app.mount("/static", StaticFiles(directory="static"), name="static")


# ---------------------------------------------------------------------------
# Admin auth (session cookie)
# ---------------------------------------------------------------------------

def require_admin(request: Request) -> str:
    """API-side admin check — returns 401 JSON when not logged in.
    When ADMIN_PASSWORD is empty, auth is disabled (prototype mode)."""
    if not ADMIN_PASSWORD:
        return "open"
    user = request.session.get("user")
    if user != ADMIN_USER:
        raise HTTPException(status_code=401, detail="not authenticated")
    return user


def is_logged_in(request: Request) -> bool:
    if not ADMIN_PASSWORD:
        return True  # auth disabled in prototype mode
    return request.session.get("user") == ADMIN_USER

# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def column_exists(conn, table: str, column: str) -> bool:
    return any(r["name"] == column for r in conn.execute(f"PRAGMA table_info({table})"))


def init_db():
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS sms_events (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                received_at TEXT    NOT NULL,
                from_number TEXT    NOT NULL,
                raw_body    TEXT    NOT NULL,
                parsed_json TEXT,
                parse_error TEXT
            );

            CREATE TABLE IF NOT EXISTS inventory (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                item        TEXT    NOT NULL,
                quantity    INTEGER NOT NULL DEFAULT 0,
                unit        TEXT    NOT NULL DEFAULT 'units',
                updated_at  TEXT    NOT NULL,
                UNIQUE(item)
            );

            CREATE TABLE IF NOT EXISTS sales_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                recorded_at TEXT    NOT NULL,
                amount_usd  REAL    NOT NULL,
                note        TEXT,
                sms_event_id INTEGER REFERENCES sms_events(id),
                site_id     INTEGER REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS item_registry (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                canonical TEXT    NOT NULL UNIQUE,
                unit      TEXT    NOT NULL DEFAULT 'units',
                aliases   TEXT    NOT NULL DEFAULT '[]',
                weight_per_unit REAL,
                degrade_per_day REAL NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS unknown_items (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                seen_at     TEXT NOT NULL,
                raw_name    TEXT NOT NULL,
                sms_event_id INTEGER REFERENCES sms_events(id),
                resolved_to TEXT,
                UNIQUE(raw_name)
            );

            CREATE TABLE IF NOT EXISTS sites (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                canonical TEXT    NOT NULL UNIQUE,
                aliases   TEXT    NOT NULL DEFAULT '[]'
            );

            CREATE TABLE IF NOT EXISTS inventory_batches (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                item         TEXT    NOT NULL,
                site_id      INTEGER REFERENCES sites(id),
                quantity     INTEGER NOT NULL,
                unit         TEXT    NOT NULL DEFAULT 'units',
                weight_lbs   REAL,
                quality_score INTEGER,
                received_at  TEXT    NOT NULL,
                sms_event_id INTEGER REFERENCES sms_events(id),
                note         TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_batches_item ON inventory_batches(item);
            CREATE INDEX IF NOT EXISTS idx_batches_received_at ON inventory_batches(received_at);
            CREATE INDEX IF NOT EXISTS idx_sales_recorded_at ON sales_log(recorded_at);

            CREATE TABLE IF NOT EXISTS attendance_log (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                recorded_at  TEXT    NOT NULL,
                site_id      INTEGER REFERENCES sites(id),
                people_count INTEGER NOT NULL,
                note         TEXT,
                sms_event_id INTEGER REFERENCES sms_events(id)
            );
            CREATE INDEX IF NOT EXISTS idx_attendance_recorded_at ON attendance_log(recorded_at);

            CREATE TABLE IF NOT EXISTS inventory_adjustments (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                recorded_at     TEXT    NOT NULL,
                item            TEXT    NOT NULL,
                kind            TEXT    NOT NULL CHECK (kind IN ('correction', 'waste', 'sold')),
                delta_quantity  INTEGER NOT NULL DEFAULT 0,
                delta_weight_lbs REAL,
                amount_usd      REAL,
                recipient       TEXT,
                note            TEXT,
                site_id         INTEGER REFERENCES sites(id),
                sales_log_id    INTEGER REFERENCES sales_log(id)
            );
            CREATE INDEX IF NOT EXISTS idx_adj_item ON inventory_adjustments(item);
            CREATE INDEX IF NOT EXISTS idx_adj_recorded_at ON inventory_adjustments(recorded_at);
        """)

        # Idempotent migrations for older DBs
        if not column_exists(conn, "item_registry", "weight_per_unit"):
            conn.execute("ALTER TABLE item_registry ADD COLUMN weight_per_unit REAL")
        if not column_exists(conn, "item_registry", "degrade_per_day"):
            conn.execute("ALTER TABLE item_registry ADD COLUMN degrade_per_day REAL NOT NULL DEFAULT 0")
        if not column_exists(conn, "sales_log", "site_id"):
            conn.execute("ALTER TABLE sales_log ADD COLUMN site_id INTEGER REFERENCES sites(id)")

    log.info("Database initialised at %s", DB_PATH)


init_db()


# ---------------------------------------------------------------------------
# Item registry helpers
# ---------------------------------------------------------------------------

# (canonical, unit, aliases, approx weight per unit in lbs, quality points lost per day)
# degrade_per_day expressed in quality-points-per-day; 0 = shelf-stable.
DEFAULT_REGISTRY = [
    ("lemon boxes",    "boxes",  ["lemons", "lmons", "lemon", "lemon box", "citrus boxes"], 30.0, 0.15),
    ("apple boxes",    "boxes",  ["apples", "apple", "apple box"], 40.0, 0.2),
    ("tomato flats",   "flats",  ["tomatoes", "tomato", "toms", "tomato flat"], 25.0, 0.4),
    ("bread loaves",   "loaves", ["bread", "loaf", "loaves", "bread loaf"], 1.0, 0.5),
    ("eggplant",       "units",  ["aubergine", "aubergines", "eggplants", "egg plant"], 1.0, 0.4),
    ("potato bags",    "bags",   ["potatoes", "potato", "spuds", "potato bag"], 10.0, 0.05),
    ("onion bags",     "bags",   ["onions", "onion", "onion bag"], 10.0, 0.05),
    ("mixed greens",   "bags",   ["greens", "salad", "salad bags", "mixed salad"], 1.0, 0.6),
    ("canned goods",   "cans",   ["cans", "tinned goods", "tins", "canned food"], 1.0, 0.0),
    ("dairy boxes",    "boxes",  ["dairy", "milk", "cheese", "dairy box"], 20.0, 0.3),
]

DEFAULT_SITES = [
    ("downtown",       ["downtown", "main", "main site", "hq"]),
    ("east side",      ["east", "east side", "eastside"]),
    ("warehouse",      ["warehouse", "wh", "depot"]),
]


def seed_registry_if_empty():
    with get_db() as conn:
        count = conn.execute("SELECT COUNT(*) FROM item_registry").fetchone()[0]
        if count == 0:
            conn.executemany(
                """INSERT OR IGNORE INTO item_registry
                   (canonical, unit, aliases, weight_per_unit, degrade_per_day)
                   VALUES (?, ?, ?, ?, ?)""",
                [(c, u, json.dumps(a), w, d) for c, u, a, w, d in DEFAULT_REGISTRY],
            )
            log.info("Seeded item registry with %d items", len(DEFAULT_REGISTRY))
        else:
            # Backfill weight_per_unit and degrade_per_day for default items that pre-date the columns
            for c, _u, _a, w, d in DEFAULT_REGISTRY:
                conn.execute(
                    "UPDATE item_registry SET weight_per_unit=? WHERE canonical=? AND weight_per_unit IS NULL",
                    (w, c),
                )
                conn.execute(
                    "UPDATE item_registry SET degrade_per_day=? WHERE canonical=? AND COALESCE(degrade_per_day, 0)=0",
                    (d, c),
                )

        site_count = conn.execute("SELECT COUNT(*) FROM sites").fetchone()[0]
        if site_count == 0:
            conn.executemany(
                "INSERT OR IGNORE INTO sites (canonical, aliases) VALUES (?, ?)",
                [(c, json.dumps(a)) for c, a in DEFAULT_SITES],
            )
            log.info("Seeded sites with %d entries", len(DEFAULT_SITES))


seed_registry_if_empty()


def backfill_batches_from_inventory():
    """One-time: if there are inventory rows but no batches, create a synthetic
    batch per item so the splash page has something to show. Idempotent."""
    with get_db() as conn:
        n_batches = conn.execute("SELECT COUNT(*) FROM inventory_batches").fetchone()[0]
        if n_batches > 0:
            return
        rows = conn.execute("SELECT item, quantity, unit, updated_at FROM inventory").fetchall()
        if not rows:
            return
        for r in rows:
            conn.execute(
                """
                INSERT INTO inventory_batches
                  (item, site_id, quantity, unit, weight_lbs, quality_score, received_at, note)
                VALUES (?, NULL, ?, ?, NULL, NULL, ?, ?)
                """,
                (r["item"], r["quantity"], r["unit"], r["updated_at"], "backfilled from inventory snapshot"),
            )
        log.info("Backfilled %d synthetic batches from inventory snapshot", len(rows))


backfill_batches_from_inventory()


def get_registry() -> list[dict]:
    with get_db() as conn:
        rows = conn.execute(
            "SELECT canonical, unit, aliases, weight_per_unit, degrade_per_day FROM item_registry ORDER BY canonical"
        ).fetchall()
    return [
        {
            "canonical": r["canonical"],
            "unit": r["unit"],
            "aliases": json.loads(r["aliases"]),
            "weight_per_unit": r["weight_per_unit"],
            "degrade_per_day": r["degrade_per_day"],
        }
        for r in rows
    ]


def get_sites() -> list[dict]:
    with get_db() as conn:
        rows = conn.execute("SELECT id, canonical, aliases FROM sites ORDER BY canonical").fetchall()
    return [
        {"id": r["id"], "canonical": r["canonical"], "aliases": json.loads(r["aliases"])}
        for r in rows
    ]


def resolve_site(name: Optional[str]) -> Optional[int]:
    """Resolve a site name (possibly an alias) to a site row id."""
    if not name:
        return None
    n = name.lower().strip()
    with get_db() as conn:
        row = conn.execute("SELECT id, aliases FROM sites WHERE canonical=?", (n,)).fetchone()
        if row:
            return row["id"]
        for r in conn.execute("SELECT id, aliases FROM sites").fetchall():
            if n in {a.lower() for a in json.loads(r["aliases"])}:
                return r["id"]
    return None


def build_parse_system() -> str:
    registry = get_registry()
    sites = get_sites()
    registry_lines = "\n".join(
        f'  - "{r["canonical"]}" ({r["unit"]}) — also accept: {", ".join(repr(a) for a in r["aliases"])}'
        for r in registry
    )
    site_lines = "\n".join(
        f'  - "{s["canonical"]}" — also accept: {", ".join(repr(a) for a in s["aliases"])}'
        for s in sites
    ) or "  (no sites registered yet)"
    return f"""
You parse text messages from food bank volunteers into structured JSON.
Messages may report:
  - sales/revenue (e.g. "sold $150", "sales 200")
  - inventory restocks (e.g. "3 lemon boxes", "got 4 apple boxes in")
  - inventory removals or usage (e.g. "used 2 bread loaves", "gave out 5 bags")
  - foot traffic / attendance (e.g. "served 80 people today", "45 visitors", "30 families came through")

Each message comes from one site. The volunteer may say the site at the start
("at downtown: 3 lemon boxes"), at the end ("3 lemon boxes — eastside"), or
omit it entirely. Resolve site names using the registry below.

Volunteers may report weight ("got 45 lbs lemons") and/or count ("3 lemon
boxes"). When weight is mentioned, capture it as weight_lbs. When omitted,
leave weight_lbs null — the server will estimate from registry weight-per-unit.

Volunteers may include a quality cue ("A grade", "great condition", "kinda
bruised", "expired"). Map to 1–5:
  5 = excellent / A grade / fresh
  4 = good
  3 = average / OK
  2 = poor / bruised / nearing expiry
  1 = bad / expired / unusable
If no quality cue, leave quality_score null.

CANONICAL ITEM REGISTRY — you MUST map any item mention to one of these exact canonical names.
Use fuzzy matching: handle typos, abbreviations, regional names (e.g. aubergine→eggplant), plurals.
{registry_lines}

SITE REGISTRY — map any site mention to one of these exact canonical names.
{site_lines}

If an item cannot be confidently matched to the registry, use the raw text as the item name
and set "unrecognized": true on that change entry so staff can review it.

Return ONLY valid JSON, no other text, matching this schema:
{{
  "site": "<canonical site name from registry, or null>",
  "sale_usd": <number or null>,
  "people_count": <integer or null — number of people / visitors / families served>,
  "changes": [
    {{
      "item": "<canonical name from registry, or raw text if unrecognized>",
      "delta": <positive=add, negative=remove>,
      "unit": "<unit>",
      "weight_lbs": <number or null>,
      "quality_score": <integer 1-5 or null>,
      "unrecognized": <true if not in registry, omit or false otherwise>
    }}
  ],
  "note": "<brief human-readable summary>"
}}

If a message is completely unrelated to inventory/sales/attendance, return:
{{"site": null, "sale_usd": null, "people_count": null, "changes": [], "note": "unrecognized message"}}
"""


# ---------------------------------------------------------------------------
# SMS parsing via Claude
# ---------------------------------------------------------------------------

def _extract_text(msg) -> str:
    """Grab the first text block from a messages-API response, even if other
    block types come first (e.g. thinking blocks)."""
    for block in msg.content:
        if getattr(block, "type", None) == "text" or hasattr(block, "text"):
            return block.text or ""
    return ""


def _extract_json(text: str) -> str:
    """Tolerantly extract a JSON object from Claude's text output.
    Strips markdown code fences and trims any preamble before the first {."""
    text = text.strip()
    if text.startswith("```"):
        # Drop the opening fence line and the closing fence
        lines = text.split("\n")
        if lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        text = "\n".join(lines[1:]).strip()
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end > start:
        return text[start:end + 1]
    return text


def parse_sms(body: str) -> dict:
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    msg = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=512,
        system=build_parse_system(),
        messages=[{"role": "user", "content": body}],
    )
    raw = _extract_text(msg)
    extracted = _extract_json(raw)
    try:
        return json.loads(extracted)
    except json.JSONDecodeError as e:
        log.error("Claude returned non-JSON. Raw response: %r", raw)
        snippet = raw[:300] if raw else "<empty response>"
        raise ValueError(f"Claude returned invalid JSON ({e}). Raw: {snippet}")


# ---------------------------------------------------------------------------
# Apply parsed update
# ---------------------------------------------------------------------------

def estimate_weight(conn, item: str, quantity: int) -> Optional[float]:
    row = conn.execute(
        "SELECT weight_per_unit FROM item_registry WHERE canonical=?", (item,)
    ).fetchone()
    if row and row["weight_per_unit"] is not None:
        return round(row["weight_per_unit"] * quantity, 2)
    return None


def apply_update(conn: sqlite3.Connection, parsed: dict, event_id: int):
    now = datetime.utcnow().isoformat()
    site_id = resolve_site(parsed.get("site"))
    note = parsed.get("note")

    if parsed.get("sale_usd"):
        conn.execute(
            "INSERT INTO sales_log (recorded_at, amount_usd, note, sms_event_id, site_id) VALUES (?, ?, ?, ?, ?)",
            (now, parsed["sale_usd"], note, event_id, site_id),
        )

    if parsed.get("people_count"):
        try:
            count = int(parsed["people_count"])
        except (TypeError, ValueError):
            count = 0
        if count > 0:
            conn.execute(
                "INSERT INTO attendance_log (recorded_at, site_id, people_count, note, sms_event_id) VALUES (?, ?, ?, ?, ?)",
                (now, site_id, count, note, event_id),
            )

    for change in parsed.get("changes", []):
        item = change["item"].lower().strip()
        delta = int(change["delta"])
        unit = change.get("unit", "units")
        weight_lbs = change.get("weight_lbs")
        quality = change.get("quality_score")

        if change.get("unrecognized"):
            conn.execute(
                "INSERT OR IGNORE INTO unknown_items (seen_at, raw_name, sms_event_id) VALUES (?, ?, ?)",
                (now, item, event_id),
            )
            log.warning("Unrecognized item in SMS %d: %r", event_id, item)
            continue

        # Restock → record a batch row
        if delta > 0:
            est_weight = weight_lbs if weight_lbs is not None else estimate_weight(conn, item, delta)
            conn.execute(
                """
                INSERT INTO inventory_batches
                  (item, site_id, quantity, unit, weight_lbs, quality_score, received_at, sms_event_id, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (item, site_id, delta, unit, est_weight, quality, now, event_id, note),
            )

        # Always update the rollup table.
        # NOTE: delta is passed twice — the VALUES clause clamps to 0 for new rows
        # (no negative initial qty), but the ON CONFLICT clause uses the raw delta
        # so removals (delta<0) actually decrement an existing row.
        conn.execute(
            """
            INSERT INTO inventory (item, quantity, unit, updated_at)
            VALUES (?, MAX(0, ?), ?, ?)
            ON CONFLICT(item) DO UPDATE SET
                quantity   = MAX(0, inventory.quantity + ?),
                unit       = excluded.unit,
                updated_at = excluded.updated_at
            """,
            (item, delta, unit, now, delta),
        )


# ---------------------------------------------------------------------------
# Twilio webhook
# ---------------------------------------------------------------------------

def _ingest_sms(from_number: str, body: str) -> dict:
    """Shared parse + persist + apply pipeline. Returns the inserted event id and parsed payload."""
    now = datetime.utcnow().isoformat()
    parsed = None
    parse_error = None
    try:
        parsed = parse_sms(body)
    except Exception as exc:
        parse_error = str(exc)
        log.error("Parse error: %s", exc)

    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO sms_events (received_at, from_number, raw_body, parsed_json, parse_error) VALUES (?, ?, ?, ?, ?)",
            (now, from_number, body, json.dumps(parsed) if parsed else None, parse_error),
        )
        event_id = cur.lastrowid

        if parsed and not parse_error:
            try:
                apply_update(conn, parsed, event_id)
            except Exception as exc:
                log.error("DB update error: %s", exc)
                parse_error = parse_error or f"apply_update: {exc}"

    return {"event_id": event_id, "parsed": parsed, "parse_error": parse_error}


@app.post("/api/simulate", dependencies=[Depends(require_admin)])
def api_simulate(payload: dict):
    """Admin-only: send a fake SMS through the parse+apply pipeline, bypassing Twilio."""
    from_number = (payload.get("from") or "+15555550100").strip()
    body = (payload.get("body") or "").strip()
    if not body:
        raise HTTPException(400, "body required")
    return _ingest_sms(from_number, body)


@app.post("/sms")
async def receive_sms(
    request: Request,
    From: str = Form(...),
    Body: str = Form(...),
):
    if VALIDATE_TWILIO:
        # In prod we either validate the Twilio signature or refuse the request.
        # An open /sms endpoint would let anyone trigger Anthropic calls + DB writes.
        if not TWILIO_AUTH_TOKEN:
            raise HTTPException(status_code=503, detail="Twilio not configured")
        validator = RequestValidator(TWILIO_AUTH_TOKEN)
        url = str(request.url)
        form_data = dict(await request.form())
        sig = request.headers.get("X-Twilio-Signature", "")
        if not validator.validate(url, form_data, sig):
            raise HTTPException(status_code=403, detail="Invalid Twilio signature")

    log.info("SMS from %s: %s", From, Body)
    _ingest_sms(From, Body)
    return HTMLResponse(
        content='<?xml version="1.0" encoding="UTF-8"?><Response></Response>',
        media_type="application/xml",
    )


# ---------------------------------------------------------------------------
# Inventory + feed + sales APIs
# ---------------------------------------------------------------------------

@app.get("/api/inventory")
def api_inventory():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT i.item, i.quantity, i.unit, i.updated_at,
                   r.weight_per_unit,
                   COALESCE(
                     (SELECT SUM(b.weight_lbs) FROM inventory_batches b WHERE b.item = i.item),
                     0
                   ) AS total_weight_lbs
            FROM inventory i
            LEFT JOIN item_registry r ON r.canonical = i.item
            ORDER BY i.item
            """
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/feed")
def api_feed(limit: int = 20):
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT id, received_at, from_number, raw_body, parsed_json, parse_error
            FROM sms_events
            ORDER BY received_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    items = []
    for r in rows:
        d = dict(r)
        d["parsed"] = json.loads(d["parsed_json"]) if d["parsed_json"] else None
        items.append(d)
    return items


@app.get("/api/sales")
def api_sales():
    with get_db() as conn:
        total = conn.execute("SELECT COALESCE(SUM(amount_usd),0) as total FROM sales_log").fetchone()["total"]
        today = conn.execute(
            "SELECT COALESCE(SUM(amount_usd),0) as total FROM sales_log WHERE DATE(recorded_at)=DATE('now')"
        ).fetchone()["total"]
        count = conn.execute("SELECT COUNT(*) as n FROM sales_log").fetchone()["n"]
    return {"total_usd": total, "today_usd": today, "transaction_count": count}


# ---------------------------------------------------------------------------
# Cashflow + totals (charts)
# ---------------------------------------------------------------------------

@app.get("/api/cashflow")
def api_cashflow(bucket: str = "day", since_days: int = 30, site: Optional[str] = None):
    """
    Returns sales aggregated by time bucket, optionally segmented by site.
    bucket: 'hour' | 'day' | 'week'
    since_days: window size in days
    site: optional canonical name filter
    """
    fmt = {
        "hour": "%Y-%m-%d %H:00",
        "day":  "%Y-%m-%d",
        "week": "%Y-W%W",
    }.get(bucket, "%Y-%m-%d")

    params = [fmt, f"-{int(since_days)} days"]
    where = ["recorded_at >= datetime('now', ?)"]
    if site:
        where.append("s.canonical = ?")
        params.append(site)
    where_sql = " AND ".join(where)

    with get_db() as conn:
        rows = conn.execute(
            f"""
            SELECT strftime(?, recorded_at) AS bucket,
                   COALESCE(s.canonical, 'unassigned') AS site,
                   SUM(l.amount_usd) AS total
            FROM sales_log l
            LEFT JOIN sites s ON s.id = l.site_id
            WHERE {where_sql}
            GROUP BY bucket, site
            ORDER BY bucket ASC
            """,
            params,
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/attendance")
def api_attendance(bucket: str = "week", since_days: int = 90, site: Optional[str] = None):
    """
    Foot traffic aggregated by time bucket, optionally segmented by site.
    bucket: 'day' | 'week' | 'month'
    """
    fmt = {
        "day":   "%Y-%m-%d",
        "week":  "%Y-W%W",
        "month": "%Y-%m",
    }.get(bucket, "%Y-W%W")

    params = [fmt, f"-{int(since_days)} days"]
    where = ["recorded_at >= datetime('now', ?)"]
    if site:
        where.append("s.canonical = ?")
        params.append(site)
    where_sql = " AND ".join(where)

    with get_db() as conn:
        rows = conn.execute(
            f"""
            SELECT strftime(?, recorded_at) AS bucket,
                   COALESCE(s.canonical, 'unassigned') AS site,
                   SUM(a.people_count) AS total
            FROM attendance_log a
            LEFT JOIN sites s ON s.id = a.site_id
            WHERE {where_sql}
            GROUP BY bucket, site
            ORDER BY bucket ASC
            """,
            params,
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/totals")
def api_totals(metric: str = "quantity", segment: str = "item"):
    """
    Aggregates current state.
    metric:  'quantity' | 'weight'
    segment: 'item' | 'site'
    """
    if segment == "site":
        with get_db() as conn:
            rows = conn.execute(
                """
                SELECT COALESCE(s.canonical, 'unassigned') AS label,
                       SUM(b.quantity) AS quantity,
                       COALESCE(SUM(b.weight_lbs), 0) AS weight
                FROM inventory_batches b
                LEFT JOIN sites s ON s.id = b.site_id
                GROUP BY label
                ORDER BY label
                """
            ).fetchall()
    else:
        with get_db() as conn:
            rows = conn.execute(
                """
                SELECT i.item AS label,
                       i.quantity AS quantity,
                       COALESCE(
                         (SELECT SUM(b.weight_lbs) FROM inventory_batches b WHERE b.item = i.item),
                         0
                       ) AS weight
                FROM inventory i
                ORDER BY i.item
                """
            ).fetchall()

    key = "weight" if metric == "weight" else "quantity"
    return [{"label": r["label"], "value": r[key] or 0} for r in rows]


# ---------------------------------------------------------------------------
# Item detail (splash page) + history
# ---------------------------------------------------------------------------

@app.get("/api/item/{canonical:path}/history")
def api_item_history(
    canonical: str,
    bucket: str = "day",
    since_days: int = 90,
    metric: str = "quantity",
):
    """
    Time-bucketed received quantity (or weight) for one item, segmented by site.
    bucket: 'day' | 'week' | 'month'
    metric: 'quantity' | 'weight'
    """
    canonical = canonical.lower().strip()
    fmt = {
        "day":   "%Y-%m-%d",
        "week":  "%Y-W%W",
        "month": "%Y-%m",
    }.get(bucket, "%Y-%m-%d")
    value_col = "SUM(b.weight_lbs)" if metric == "weight" else "SUM(b.quantity)"

    with get_db() as conn:
        rows = conn.execute(
            f"""
            SELECT strftime(?, b.received_at) AS bucket,
                   COALESCE(s.canonical, 'unassigned') AS site,
                   COALESCE({value_col}, 0) AS total
            FROM inventory_batches b
            LEFT JOIN sites s ON s.id = b.site_id
            WHERE b.item = ?
              AND b.received_at >= datetime('now', ?)
            GROUP BY bucket, site
            ORDER BY bucket ASC
            """,
            (fmt, canonical, f"-{int(since_days)} days"),
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/item/{canonical:path}/adjustments")
def api_item_adjustments(canonical: str):
    canonical = canonical.lower().strip()
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT a.id, a.recorded_at, a.kind, a.delta_quantity, a.delta_weight_lbs,
                   a.amount_usd, a.recipient, a.note,
                   COALESCE(s.canonical, 'unassigned') AS site
            FROM inventory_adjustments a
            LEFT JOIN sites s ON s.id = a.site_id
            WHERE a.item = ?
            ORDER BY a.recorded_at DESC
            """,
            (canonical,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/item/{canonical:path}")
def api_item_detail(canonical: str):
    canonical = canonical.lower().strip()
    with get_db() as conn:
        reg = conn.execute(
            "SELECT canonical, unit, aliases, weight_per_unit, degrade_per_day FROM item_registry WHERE canonical=?",
            (canonical,),
        ).fetchone()
        inv = conn.execute(
            "SELECT quantity, unit, updated_at FROM inventory WHERE item=?",
            (canonical,),
        ).fetchone()
        batches = conn.execute(
            """
            SELECT b.id, b.quantity, b.unit, b.weight_lbs, b.quality_score,
                   b.received_at, b.note,
                   COALESCE(s.canonical, 'unassigned') AS site
            FROM inventory_batches b
            LEFT JOIN sites s ON s.id = b.site_id
            WHERE b.item = ?
            ORDER BY b.received_at DESC
            """,
            (canonical,),
        ).fetchall()

    if not reg and not inv and not batches:
        raise HTTPException(404, f"item '{canonical}' not found")

    degrade = (reg["degrade_per_day"] or 0) if reg else 0
    now = datetime.utcnow()

    batches_list = []
    for b in batches:
        d = dict(b)
        # Compute current degraded quality if we have a starting score and a degrade rate
        if d["quality_score"] is not None:
            received = datetime.fromisoformat(d["received_at"])
            days_old = max(0.0, (now - received).total_seconds() / 86400.0)
            d["days_old"] = round(days_old, 2)
            current = d["quality_score"] - degrade * days_old
            d["current_quality"] = round(max(0.0, current), 2)
        else:
            d["days_old"] = None
            d["current_quality"] = None
        batches_list.append(d)

    current_quality_vals = [b["current_quality"] for b in batches_list if b["current_quality"] is not None]
    initial_quality_vals = [b["quality_score"]   for b in batches_list if b["quality_score"]   is not None]
    avg_current_quality = round(sum(current_quality_vals) / len(current_quality_vals), 2) if current_quality_vals else None
    avg_initial_quality = round(sum(initial_quality_vals) / len(initial_quality_vals), 2) if initial_quality_vals else None
    total_weight = round(sum((b["weight_lbs"] or 0) for b in batches_list), 2)

    return {
        "canonical": canonical,
        "registry": dict(reg) | {"aliases": json.loads(reg["aliases"])} if reg else None,
        "current": dict(inv) if inv else None,
        "batches": batches_list,
        "avg_quality": avg_current_quality,            # current (degraded)
        "avg_initial_quality": avg_initial_quality,    # as-received
        "total_weight_lbs": total_weight,
    }


# ---------------------------------------------------------------------------
# Registry + sites + unknowns management
# ---------------------------------------------------------------------------

@app.get("/api/registry")
def api_registry():
    return get_registry()


@app.post("/api/registry", dependencies=[Depends(require_admin)])
def api_registry_add(payload: dict):
    canonical = payload.get("canonical", "").lower().strip()
    unit = payload.get("unit", "units")
    aliases = payload.get("aliases", [])
    weight_per_unit = payload.get("weight_per_unit")
    degrade_per_day = payload.get("degrade_per_day") or 0
    if not canonical:
        raise HTTPException(400, "canonical required")
    with get_db() as conn:
        conn.execute(
            """INSERT INTO item_registry
               (canonical, unit, aliases, weight_per_unit, degrade_per_day)
               VALUES (?, ?, ?, ?, ?)""",
            (canonical, unit, json.dumps(aliases), weight_per_unit, degrade_per_day),
        )
    return {"ok": True, "canonical": canonical}


@app.patch("/api/registry/{canonical:path}", dependencies=[Depends(require_admin)])
def api_registry_update(canonical: str, payload: dict):
    """Update an existing item: aliases (append), unit, weight_per_unit, degrade_per_day."""
    new_aliases = payload.get("aliases")
    unit = payload.get("unit")
    weight_per_unit = payload.get("weight_per_unit")
    degrade_per_day = payload.get("degrade_per_day")
    with get_db() as conn:
        row = conn.execute(
            "SELECT aliases FROM item_registry WHERE canonical=?", (canonical,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "item not found")
        merged = json.loads(row["aliases"])
        if new_aliases:
            merged = list(dict.fromkeys(merged + new_aliases))
        conn.execute(
            """UPDATE item_registry
               SET aliases=?,
                   unit=COALESCE(?, unit),
                   weight_per_unit=COALESCE(?, weight_per_unit),
                   degrade_per_day=COALESCE(?, degrade_per_day)
               WHERE canonical=?""",
            (json.dumps(merged), unit, weight_per_unit, degrade_per_day, canonical),
        )
    return {"ok": True, "canonical": canonical, "aliases": merged}


@app.delete("/api/registry/{canonical:path}", dependencies=[Depends(require_admin)])
def api_registry_delete(canonical: str):
    with get_db() as conn:
        conn.execute("DELETE FROM item_registry WHERE canonical=?", (canonical,))
    return {"ok": True}


@app.get("/api/sites")
def api_sites():
    return get_sites()


@app.post("/api/sites", dependencies=[Depends(require_admin)])
def api_sites_add(payload: dict):
    canonical = payload.get("canonical", "").lower().strip()
    aliases = payload.get("aliases", [])
    if not canonical:
        raise HTTPException(400, "canonical required")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO sites (canonical, aliases) VALUES (?, ?)",
            (canonical, json.dumps(aliases)),
        )
    return {"ok": True, "canonical": canonical}


@app.patch("/api/sites/{canonical:path}", dependencies=[Depends(require_admin)])
def api_sites_update(canonical: str, payload: dict):
    new_aliases = payload.get("aliases", [])
    with get_db() as conn:
        row = conn.execute("SELECT aliases FROM sites WHERE canonical=?", (canonical,)).fetchone()
        if not row:
            raise HTTPException(404, "site not found")
        merged = list(dict.fromkeys(json.loads(row["aliases"]) + new_aliases))
        conn.execute("UPDATE sites SET aliases=? WHERE canonical=?", (json.dumps(merged), canonical))
    return {"ok": True, "canonical": canonical, "aliases": merged}


@app.delete("/api/sites/{canonical:path}", dependencies=[Depends(require_admin)])
def api_sites_delete(canonical: str):
    with get_db() as conn:
        conn.execute("DELETE FROM sites WHERE canonical=?", (canonical,))
    return {"ok": True}


@app.get("/api/unknowns", dependencies=[Depends(require_admin)])
def api_unknowns():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT u.id, u.seen_at, u.raw_name, u.resolved_to, u.sms_event_id, e.raw_body
            FROM unknown_items u
            LEFT JOIN sms_events e ON e.id = u.sms_event_id
            WHERE u.resolved_to IS NULL
            ORDER BY u.seen_at DESC
            """
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/unknowns/{unknown_id}/resolve", dependencies=[Depends(require_admin)])
def api_resolve_unknown(unknown_id: int, payload: dict):
    action = payload.get("action")
    canonical = payload.get("canonical", "").lower().strip()
    if not canonical:
        raise HTTPException(400, "canonical required")

    with get_db() as conn:
        row = conn.execute(
            "SELECT raw_name, sms_event_id, resolved_to FROM unknown_items WHERE id=?",
            (unknown_id,),
        ).fetchone()
        if not row:
            raise HTTPException(404, "unknown item not found")
        if row["resolved_to"]:
            raise HTTPException(400, f"already resolved to '{row['resolved_to']}'")
        raw_name = row["raw_name"]
        event_id = row["sms_event_id"]

        if action == "add":
            unit = payload.get("unit", "units")
            weight_per_unit = payload.get("weight_per_unit")
            degrade_per_day = payload.get("degrade_per_day") or 0
            conn.execute(
                """INSERT OR IGNORE INTO item_registry
                   (canonical, unit, aliases, weight_per_unit, degrade_per_day)
                   VALUES (?, ?, ?, ?, ?)""",
                (canonical, unit, json.dumps([raw_name]), weight_per_unit, degrade_per_day),
            )
        elif action == "map":
            alias_row = conn.execute(
                "SELECT aliases FROM item_registry WHERE canonical=?", (canonical,)
            ).fetchone()
            if not alias_row:
                raise HTTPException(404, f"canonical '{canonical}' not in registry")
            existing = json.loads(alias_row["aliases"])
            if raw_name not in existing:
                existing.append(raw_name)
                conn.execute(
                    "UPDATE item_registry SET aliases=? WHERE canonical=?",
                    (json.dumps(existing), canonical),
                )
        else:
            raise HTTPException(400, "action must be 'map' or 'add'")

        # Replay the deferred inventory change from the originating SMS
        replayed = []
        if event_id is not None:
            evt = conn.execute(
                "SELECT parsed_json FROM sms_events WHERE id=?", (event_id,)
            ).fetchone()
            if evt and evt["parsed_json"]:
                parsed = json.loads(evt["parsed_json"])
                recovered = [
                    {**ch, "item": canonical, "unrecognized": False}
                    for ch in parsed.get("changes", [])
                    if ch.get("unrecognized") and (ch.get("item") or "").lower().strip() == raw_name
                ]
                if recovered:
                    apply_update(
                        conn,
                        {**parsed, "changes": recovered},
                        event_id,
                    )
                    replayed = recovered

        conn.execute(
            "UPDATE unknown_items SET resolved_to=? WHERE id=?", (canonical, unknown_id)
        )

    return {"ok": True, "raw_name": raw_name, "resolved_to": canonical, "replayed": replayed}


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

@app.get("/api/reconcile/summary")
def api_reconcile_summary():
    """Per-site rollups of sales + attendance, plus counts of unassigned rows."""
    with get_db() as conn:
        sales_per_site = conn.execute(
            """
            SELECT COALESCE(s.canonical, 'unassigned') AS site,
                   COUNT(*) AS count,
                   COALESCE(SUM(amount_usd), 0) AS total_usd
            FROM sales_log l
            LEFT JOIN sites s ON s.id = l.site_id
            GROUP BY site
            ORDER BY site
            """
        ).fetchall()
        attendance_per_site = conn.execute(
            """
            SELECT COALESCE(s.canonical, 'unassigned') AS site,
                   COUNT(*) AS count,
                   COALESCE(SUM(people_count), 0) AS total_people
            FROM attendance_log a
            LEFT JOIN sites s ON s.id = a.site_id
            GROUP BY site
            ORDER BY site
            """
        ).fetchall()
        unassigned_sales = conn.execute(
            """
            SELECT l.id, l.recorded_at, l.amount_usd, l.note,
                   e.raw_body
            FROM sales_log l
            LEFT JOIN sms_events e ON e.id = l.sms_event_id
            WHERE l.site_id IS NULL
            ORDER BY l.recorded_at DESC
            """
        ).fetchall()
        unassigned_attendance = conn.execute(
            """
            SELECT a.id, a.recorded_at, a.people_count, a.note,
                   e.raw_body
            FROM attendance_log a
            LEFT JOIN sms_events e ON e.id = a.sms_event_id
            WHERE a.site_id IS NULL
            ORDER BY a.recorded_at DESC
            """
        ).fetchall()

    return {
        "sales_per_site":      [dict(r) for r in sales_per_site],
        "attendance_per_site": [dict(r) for r in attendance_per_site],
        "unassigned_sales":      [dict(r) for r in unassigned_sales],
        "unassigned_attendance": [dict(r) for r in unassigned_attendance],
    }


@app.post("/api/sales/{sale_id}/assign-site", dependencies=[Depends(require_admin)])
def api_assign_sale_site(sale_id: int, payload: dict):
    site_name = (payload.get("site") or "").strip()
    site_id = resolve_site(site_name)
    if site_name and not site_id:
        raise HTTPException(400, f"site '{site_name}' not in registry")
    with get_db() as conn:
        r = conn.execute("UPDATE sales_log SET site_id=? WHERE id=?", (site_id, sale_id))
        if r.rowcount == 0:
            raise HTTPException(404, "sale not found")
    return {"ok": True, "sale_id": sale_id, "site": site_name or None}


@app.post("/api/attendance/{att_id}/assign-site", dependencies=[Depends(require_admin)])
def api_assign_attendance_site(att_id: int, payload: dict):
    site_name = (payload.get("site") or "").strip()
    site_id = resolve_site(site_name)
    if site_name and not site_id:
        raise HTTPException(400, f"site '{site_name}' not in registry")
    with get_db() as conn:
        r = conn.execute("UPDATE attendance_log SET site_id=? WHERE id=?", (site_id, att_id))
        if r.rowcount == 0:
            raise HTTPException(404, "attendance not found")
    return {"ok": True, "attendance_id": att_id, "site": site_name or None}


# ---------------------------------------------------------------------------
# Per-item adjustments (correction / waste / sold)
# ---------------------------------------------------------------------------

VALID_ADJ_KINDS = {"correction", "waste", "sold"}


@app.post("/api/item/{canonical:path}/adjust", dependencies=[Depends(require_admin)])
def api_item_adjust(canonical: str, payload: dict):
    """
    Apply an adjustment to an item's inventory and log it for audit.
    Body: {kind, delta_quantity, delta_weight_lbs?, amount_usd?, recipient?, note?, site?}
    Kinds: 'correction' (delta can be +/-), 'waste' (delta forced negative),
           'sold' (delta forced negative; if amount_usd>0, also writes a sales_log row).
    """
    canonical = canonical.lower().strip()
    kind = (payload.get("kind") or "").strip().lower()
    if kind not in VALID_ADJ_KINDS:
        raise HTTPException(400, f"kind must be one of {sorted(VALID_ADJ_KINDS)}")

    try:
        delta_q = int(payload.get("delta_quantity") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "delta_quantity must be an integer")

    delta_w = payload.get("delta_weight_lbs")
    if delta_w in ("", None):
        delta_w = None
    else:
        try:
            delta_w = float(delta_w)
        except (TypeError, ValueError):
            raise HTTPException(400, "delta_weight_lbs must be numeric")

    amount = payload.get("amount_usd")
    if amount in ("", None):
        amount = None
    else:
        try:
            amount = float(amount)
        except (TypeError, ValueError):
            raise HTTPException(400, "amount_usd must be numeric")

    recipient = (payload.get("recipient") or "").strip() or None
    note      = (payload.get("note") or "").strip() or None
    site_name = (payload.get("site") or "").strip() or None
    site_id   = resolve_site(site_name) if site_name else None

    # Force sign by kind so the UI can't accidentally invert a "waste" or "sold"
    if kind in ("waste", "sold") and delta_q > 0:
        delta_q = -delta_q
    if kind in ("waste", "sold") and delta_w is not None and delta_w > 0:
        delta_w = -delta_w

    if delta_q == 0 and not (kind == "sold" and amount):
        raise HTTPException(400, "delta_quantity required (or for 'sold', an amount_usd)")

    now = datetime.utcnow().isoformat()
    with get_db() as conn:
        sales_id = None
        if kind == "sold" and amount and amount > 0:
            cur = conn.execute(
                "INSERT INTO sales_log (recorded_at, amount_usd, note, site_id) VALUES (?, ?, ?, ?)",
                (now, amount, f"sold {canonical}" + (f" to {recipient}" if recipient else ""), site_id),
            )
            sales_id = cur.lastrowid

        cur = conn.execute(
            """INSERT INTO inventory_adjustments
                 (recorded_at, item, kind, delta_quantity, delta_weight_lbs, amount_usd, recipient, note, site_id, sales_log_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (now, canonical, kind, delta_q, delta_w, amount, recipient, note, site_id, sales_id),
        )
        adj_id = cur.lastrowid

        if delta_q != 0:
            conn.execute(
                """
                INSERT INTO inventory (item, quantity, unit, updated_at)
                VALUES (?, MAX(0, ?), 'units', ?)
                ON CONFLICT(item) DO UPDATE SET
                    quantity   = MAX(0, inventory.quantity + ?),
                    updated_at = excluded.updated_at
                """,
                (canonical, delta_q, now, delta_q),
            )

    return {"ok": True, "id": adj_id, "kind": kind, "delta_quantity": delta_q, "sales_log_id": sales_id}


# ---------------------------------------------------------------------------
# Spreadsheet import (Haymarket-style weekly worksheet)
# ---------------------------------------------------------------------------
import re


def _parse_qty(raw) -> tuple[Optional[int], str]:
    """'3 boxes' -> (3, 'boxes'); 60 -> (60, 'units'); None -> (None, 'units')."""
    if raw is None:
        return None, "units"
    if isinstance(raw, (int, float)):
        return int(raw), "units"
    s = str(raw).strip()
    m = re.match(r"\s*(\d+(?:\.\d+)?)\s*(.*)$", s)
    if not m:
        return None, "units"
    qty = int(float(m.group(1)))
    unit = (m.group(2).strip() or "units").lower()
    return qty, unit


def _extract_money(text: str) -> Optional[float]:
    """Find a dollar amount in a free-text cell. Returns float or None.
    Filters out per-unit rates like '.77/lb'."""
    if not isinstance(text, str):
        return None
    if re.search(r"/\s*(lb|lbs|oz|kg|pound|each|count)\b", text, re.I):
        return None
    m = re.search(r"\$?\s*(\d+(?:\.\d+)?)", text)
    if m:
        return float(m.group(1))
    return None


def parse_haymarket_xlsx(file_bytes: bytes) -> dict:
    """Parse a Haymarket-style weekly worksheet into a preview structure."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("worksheet has fewer than 2 rows")

    # --- Date ---
    date_iso = None
    for cell in rows[0]:
        if isinstance(cell, str) and "DATE" in cell.upper():
            m = re.search(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", cell)
            if m:
                mo, dy, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if yr < 100:
                    yr += 2000
                date_iso = f"{yr:04d}-{mo:02d}-{dy:02d}"
                break

    # --- Detect sections by walking row 1 ---
    # Section 0 is always col 0 (the implicit source like "Haymarket").
    # Subsequent sections start at columns where row1 says "for distribution".
    row0 = list(rows[0]) + [None] * 8
    row1 = list(rows[1]) + [None] * 8
    DISTRIB_HEADERS = {"for distribution", "for distribiution", "distribution"}
    SKIP_NAMES = {"total", "compost", "free", "goal", ""}

    sections = []
    # Section 0 — acquisition source
    src_label = row1[0].strip() if isinstance(row1[0], str) and row1[0].strip() else "Source"
    sections.append({"label": src_label, "start_col": 0})
    for col in range(1, len(row1)):
        sub = row1[col]
        if isinstance(sub, str) and sub.strip().lower() in DISTRIB_HEADERS:
            label = row0[col].strip() if isinstance(row0[col], str) and row0[col].strip() else f"col_{col}"
            sections.append({"label": label, "start_col": col})
    for i, sec in enumerate(sections):
        end = sections[i + 1]["start_col"] if i + 1 < len(sections) else len(row0)
        sec["width"] = end - sec["start_col"]

    # --- Walk rows in each section ---
    out_sections = []
    for sec in sections:
        sc = sec["start_col"]
        width = sec["width"]
        is_distribution = sc != 0  # col 0 is always source; others are "for distribution"
        items = []
        raw_text = []
        for r_idx in range(2, len(rows)):
            cells = list(rows[r_idx][sc:sc + width])
            if not any(cells):
                continue
            item_name = cells[0] if len(cells) > 0 else None
            qty_raw = cells[1] if len(cells) > 1 else None
            lbs = cells[2] if len(cells) > 2 else None
            name_ok = isinstance(item_name, str) and item_name.strip().lower() not in SKIP_NAMES
            qty_ok = isinstance(qty_raw, (int, float)) or (isinstance(qty_raw, str) and re.search(r"\d", qty_raw))
            lbs_ok = isinstance(lbs, (int, float))

            # Acquisition cols: an item is anything with a name and qty-or-lbs.
            # Distribution cols: require ALL three (name + qty + lbs) — that's what
            # YMCA-style "for distribution but really an inventory list" looks like.
            # Free-text in distribution cols (e.g. "VA sold 989", ".77/lb") goes to raw_text.
            if is_distribution:
                is_item = name_ok and qty_ok and lbs_ok
            else:
                is_item = name_ok and (qty_ok or lbs_ok)

            if is_item:
                qty, unit = _parse_qty(qty_raw)
                items.append({
                    "item": item_name.strip().lower(),
                    "quantity": qty,
                    "unit": unit,
                    "weight_lbs": float(lbs) if isinstance(lbs, (int, float)) else None,
                    "row": r_idx + 1,
                })
            else:
                txt_parts = [str(c).strip() for c in cells if c is not None and str(c).strip()]
                if txt_parts:
                    raw_text.append(" ".join(txt_parts))

        # Decide section type
        if not is_distribution:
            stype = "acquisition"
        elif items:
            stype = "ambiguous"  # like YMCA — distribution header but full item rows
        else:
            stype = "distribution"

        # For pure distribution sections, scrape $ amounts from raw_text
        sales = []
        if stype == "distribution":
            for txt in raw_text:
                amt = _extract_money(txt)
                if amt is not None:
                    sales.append({"raw": txt, "amount_usd": amt})

        out_sections.append({
            "label": sec["label"],
            "type": stype,
            "items": items,
            "sales": sales,
            "raw_text": raw_text,
        })

    # Suggest registry mappings for each acquisition/ambiguous item
    registry = {r["canonical"]: r for r in get_registry()}
    alias_index = {}
    for r in registry.values():
        for a in r["aliases"]:
            alias_index[a.lower()] = r["canonical"]
        alias_index[r["canonical"].lower()] = r["canonical"]
    for sec in out_sections:
        for it in sec["items"]:
            it["suggested_canonical"] = alias_index.get(it["item"]) or it["item"]
            it["in_registry"] = it["suggested_canonical"] in registry

    return {"date": date_iso, "sections": out_sections}


@app.post("/api/import/preview", dependencies=[Depends(require_admin)])
async def api_import_preview(file: UploadFile = File(...)):
    content = await file.read()
    try:
        preview = parse_haymarket_xlsx(content)
    except Exception as exc:
        raise HTTPException(400, f"could not parse xlsx: {exc}")
    preview["filename"] = file.filename
    return preview


@app.post("/api/import/commit", dependencies=[Depends(require_admin)])
def api_import_commit(payload: dict):
    """
    Body shape (echoed from the preview, with user edits):
    {
      "date": "2026-05-16",
      "filename": "...",
      "sections": [
        {"label": "Haymarket", "type": "acquisition", "items": [{item, quantity, unit, weight_lbs, suggested_canonical, skip?}], "sales": [], "raw_text": []},
        {"label": "Archdale", "type": "distribution", "items": [], "sales": [{raw, amount_usd, site_override?}], ...},
        ...
      ]
    }
    """
    date = payload.get("date")
    if not date:
        raise HTTPException(400, "date required")
    received_at = f"{date}T00:00:00"

    created_batches = 0
    created_sales = 0
    new_sites = []
    new_items = []

    with get_db() as conn:
        for sec in payload.get("sections", []):
            section_site_id = resolve_site(sec.get("label"))
            if sec["type"] in ("acquisition", "ambiguous"):
                # For ambiguous (like YMCA), the UI sets per-item skip flags if user wants to skip
                for it in sec.get("items", []):
                    if it.get("skip"):
                        continue
                    canonical = (it.get("suggested_canonical") or it.get("item") or "").lower().strip()
                    if not canonical:
                        continue
                    # Auto-register if new
                    reg = conn.execute(
                        "SELECT canonical FROM item_registry WHERE canonical=?", (canonical,)
                    ).fetchone()
                    if not reg:
                        unit = (it.get("unit") or "units").lower()
                        wpu = None
                        qty = it.get("quantity") or 0
                        lbs = it.get("weight_lbs")
                        if qty and lbs:
                            wpu = round(lbs / qty, 2)
                        conn.execute(
                            "INSERT OR IGNORE INTO item_registry (canonical, unit, aliases, weight_per_unit) VALUES (?, ?, ?, ?)",
                            (canonical, unit, json.dumps([it.get("item", canonical)]), wpu),
                        )
                        new_items.append(canonical)
                    # Write the batch
                    conn.execute(
                        """INSERT INTO inventory_batches
                             (item, site_id, quantity, unit, weight_lbs, quality_score, received_at, note)
                           VALUES (?, ?, ?, ?, ?, NULL, ?, ?)""",
                        (
                            canonical,
                            section_site_id,
                            int(it.get("quantity") or 0),
                            (it.get("unit") or "units"),
                            float(it["weight_lbs"]) if it.get("weight_lbs") is not None else None,
                            received_at,
                            f"imported from {payload.get('filename', 'xlsx')} ({sec['label']})",
                        ),
                    )
                    conn.execute(
                        """INSERT INTO inventory (item, quantity, unit, updated_at)
                           VALUES (?, MAX(0, ?), ?, ?)
                           ON CONFLICT(item) DO UPDATE SET
                               quantity   = MAX(0, inventory.quantity + ?),
                               unit       = excluded.unit,
                               updated_at = excluded.updated_at""",
                        (canonical, int(it.get("quantity") or 0), (it.get("unit") or "units"),
                         received_at, int(it.get("quantity") or 0)),
                    )
                    created_batches += 1
            elif sec["type"] == "distribution":
                # Auto-register distribution sites if new
                if not section_site_id and sec.get("label"):
                    site_name = sec["label"].lower().strip()
                    conn.execute(
                        "INSERT OR IGNORE INTO sites (canonical, aliases) VALUES (?, ?)",
                        (site_name, json.dumps([])),
                    )
                    section_site_id = resolve_site(site_name)
                    new_sites.append(site_name)
                for sale in sec.get("sales", []):
                    if sale.get("skip"):
                        continue
                    amt = sale.get("amount_usd")
                    if not amt:
                        continue
                    conn.execute(
                        "INSERT INTO sales_log (recorded_at, amount_usd, note, site_id) VALUES (?, ?, ?, ?)",
                        (received_at, float(amt),
                         f"imported from {payload.get('filename', 'xlsx')}: {sale.get('raw', '')}",
                         section_site_id),
                    )
                    created_sales += 1

    return {
        "ok": True,
        "created_batches": created_batches,
        "created_sales": created_sales,
        "new_items": new_items,
        "new_sites": new_sites,
    }


# ---------------------------------------------------------------------------
# Row-level list endpoints (for the dashboard tabs)
# ---------------------------------------------------------------------------

@app.get("/api/sales/list")
def api_sales_list(limit: int = 50):
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT l.id, l.recorded_at, l.amount_usd,
                   COALESCE(s.canonical, 'unassigned') AS site,
                   l.note, l.sms_event_id
            FROM sales_log l
            LEFT JOIN sites s ON s.id = l.site_id
            ORDER BY l.recorded_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/attendance/list")
def api_attendance_list(limit: int = 50):
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT a.id, a.recorded_at,
                   COALESCE(s.canonical, 'unassigned') AS site,
                   a.people_count, a.note, a.sms_event_id
            FROM attendance_log a
            LEFT JOIN sites s ON s.id = a.site_id
            ORDER BY a.recorded_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/adjustments")
def api_adjustments_list(limit: int = 50):
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT a.id, a.recorded_at, a.item, a.kind, a.delta_quantity,
                   a.delta_weight_lbs, a.amount_usd, a.recipient, a.note,
                   COALESCE(s.canonical, 'unassigned') AS site
            FROM inventory_adjustments a
            LEFT JOIN sites s ON s.id = a.site_id
            ORDER BY a.recorded_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.get("/api/batches")
def api_batches_list(limit: int = 50):
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT b.id, b.item, COALESCE(s.canonical, 'unassigned') AS site,
                   b.quantity, b.unit, b.weight_lbs, b.quality_score,
                   b.received_at, b.note
            FROM inventory_batches b
            LEFT JOIN sites s ON s.id = b.site_id
            ORDER BY b.received_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# CSV downloads
# ---------------------------------------------------------------------------

def _csv_response(rows: list[dict], filename: str) -> Response:
    if not rows:
        text = ""
    else:
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
        text = buf.getvalue()
    return Response(
        content=text,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/inventory.csv")
def api_inventory_csv():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT i.item, i.quantity, i.unit, i.updated_at,
                   r.weight_per_unit,
                   r.degrade_per_day,
                   COALESCE((SELECT SUM(b.weight_lbs) FROM inventory_batches b WHERE b.item=i.item), 0) AS total_weight_lbs
            FROM inventory i
            LEFT JOIN item_registry r ON r.canonical = i.item
            ORDER BY i.item
            """
        ).fetchall()
    return _csv_response([dict(r) for r in rows], "inventory.csv")


@app.get("/api/batches.csv")
def api_batches_csv(item: Optional[str] = None):
    sql = """
        SELECT b.id, b.item, COALESCE(s.canonical, 'unassigned') AS site,
               b.quantity, b.unit, b.weight_lbs, b.quality_score,
               b.received_at, b.note
        FROM inventory_batches b
        LEFT JOIN sites s ON s.id = b.site_id
    """
    params = []
    if item:
        sql += " WHERE b.item = ?"
        params.append(item.lower().strip())
    sql += " ORDER BY b.received_at DESC"
    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
    fname = f"batches-{item}.csv" if item else "batches.csv"
    return _csv_response([dict(r) for r in rows], fname)


@app.get("/api/sales.csv")
def api_sales_csv():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT l.id, l.recorded_at, l.amount_usd,
                   COALESCE(s.canonical, 'unassigned') AS site,
                   l.note, l.sms_event_id
            FROM sales_log l
            LEFT JOIN sites s ON s.id = l.site_id
            ORDER BY l.recorded_at DESC
            """
        ).fetchall()
    return _csv_response([dict(r) for r in rows], "sales.csv")


@app.get("/api/attendance.csv")
def api_attendance_csv():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT a.id, a.recorded_at, COALESCE(s.canonical, 'unassigned') AS site,
                   a.people_count, a.note, a.sms_event_id
            FROM attendance_log a
            LEFT JOIN sites s ON s.id = a.site_id
            ORDER BY a.recorded_at DESC
            """
        ).fetchall()
    return _csv_response([dict(r) for r in rows], "attendance.csv")


@app.get("/api/adjustments.csv")
def api_adjustments_csv(item: Optional[str] = None):
    sql = """
        SELECT id, recorded_at, item, kind, delta_quantity, delta_weight_lbs,
               amount_usd, recipient, note
        FROM inventory_adjustments
    """
    params = []
    if item:
        sql += " WHERE item = ?"
        params.append(item.lower().strip())
    sql += " ORDER BY recorded_at DESC"
    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
    fname = f"adjustments-{item}.csv" if item else "adjustments.csv"
    return _csv_response([dict(r) for r in rows], fname)


# ---------------------------------------------------------------------------
# HTML pages
# ---------------------------------------------------------------------------

def _render(path: str) -> str:
    with open(path) as f:
        return f.read()


@app.get("/", response_class=HTMLResponse)
def dashboard():
    return _render("templates/index.html")


@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/login?next=/admin", status_code=303)
    return HTMLResponse(_render("templates/admin.html"))


@app.get("/simulator", response_class=HTMLResponse)
def simulator_page(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/login?next=/simulator", status_code=303)
    return HTMLResponse(_render("templates/simulator.html"))


@app.get("/reconcile", response_class=HTMLResponse)
def reconcile_page(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/login?next=/reconcile", status_code=303)
    return HTMLResponse(_render("templates/reconcile.html"))


@app.get("/history", response_class=HTMLResponse)
def history_page():
    return HTMLResponse(_render("templates/history.html"))


@app.get("/inventory", response_class=HTMLResponse)
def inventory_page():
    return HTMLResponse(_render("templates/inventory.html"))


@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/login?next=/import", status_code=303)
    return HTMLResponse(_render("templates/import.html"))


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, next: str = "/admin"):
    if is_logged_in(request):
        return RedirectResponse(url=next, status_code=303)
    return HTMLResponse(_render("templates/login.html"))


@app.post("/login")
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    next: str = Form("/admin"),
):
    if not ADMIN_PASSWORD:
        raise HTTPException(status_code=503, detail="Admin auth not configured")
    ok_user = secrets.compare_digest(username, ADMIN_USER)
    ok_pass = secrets.compare_digest(password, ADMIN_PASSWORD)
    if not (ok_user and ok_pass):
        # Re-render the login page with an error flag
        html = _render("templates/login.html").replace(
            "<!--ERROR-->", '<p class="error">Invalid username or password.</p>'
        )
        return HTMLResponse(html, status_code=401)
    request.session["user"] = ADMIN_USER
    # Only allow same-origin redirects
    safe_next = next if next.startswith("/") and not next.startswith("//") else "/admin"
    return RedirectResponse(url=safe_next, status_code=303)


@app.post("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


@app.get("/item/{canonical:path}", response_class=HTMLResponse)
def item_page(canonical: str):
    return _render("templates/item.html")
